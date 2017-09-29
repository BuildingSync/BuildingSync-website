
import xml.etree.ElementTree as et
from xml.dom import minidom

import openpyxl

# Known limitations:
# 1) Some of the keys are not scrubbed for those asterisks
# 2) The get*** functions are vulnerable to data loss if keys are not unique
# 3) Order is not preserved in tables, need to use the newer getlistinfo function
# 4) Messing with the labels was probably a mistake
# 5) The Energy Sources table is broken
# 6) There are likely missing refs between the building and measures

energysources_labels = ['Energy Source',
                        'ID',
                        None,
                        None,
                        'Type',
                        'Rate schedule']

spacefunctions_211_labels = ['Space Number',
                             'Function type*',
                             'Original intended use',
                             '''Gross Floor Area*
(per space)''',
                             '''Conditioned Area*
(Approx % of total function space)''',
                             'Number of Occupants',
                             'Approximate Plug Loads (W/sf)',
                             'Use (hours/week)',
                             'Use (weeks/year)',
                             'Principal HVAC Type*',
                             'Principal Lighting Type*']

spacefunctions_labels = ['Space Number',
                         'Function type',
                         'Original intended use',
                         'Gross Floor Area',
                         'Percent Conditioned Area',
                         'Number of Occupants',
                         'Approximate Plug Loads (W/sf)',
                         'Use (hours/week)',
                         'Use (weeks/year)',
                         'Principal HVAC Type',
                         'Principal Lighting Type']

eemsummary_header_yi = ['Low-Cost and No-Cost Recommendations',
                        'Modified System',
                        'Impact on Occupant Comfort or IEQ',
                        'Other Non-Energy Impacts',
                        'Cost',
                        'Savings Impact',
                        'Typical ROI',
                        'Priority']

eemsummary_header_er = ['Potential Capital Recommendations',
                        'Modified System',
                        'Impact on Occupant Comfort',
                        'Other Non-Energy Impacts',
                        'Cost',
                        'Savings Impact',
                        'Typical ROI',
                        'Priority']


def tuple_from_coordinate(coordinate):
    coord = openpyxl.utils.coordinate_from_string(coordinate)
    col = openpyxl.utils.column_index_from_string(coord[0])
    row = coord[1]
    return (col, row)


def reportSpaceFunctionsLabelDiff(labels):
    print('Space functions label difference!')


class ScanFailure(Exception):
    pass


class LabelMismatch(Exception):
    pass


class MissingRequired(Exception):
    pass


def cellrange(worksheet, mincol=None, minrow=None, maxcol=None, maxrow=None):
    if minrow == maxrow:
        for row in worksheet.iter_rows(min_row=minrow, min_col=mincol,
                                       max_col=maxcol, max_row=maxrow):
            return [el.value for el in row]
    elif mincol == maxcol:
        for col in worksheet.iter_cols(min_row=minrow, min_col=mincol,
                                       max_col=maxcol, max_row=maxrow):
            return [el.value for el in col]
    results = []
    for row in worksheet.iter_rows(min_row=minrow, min_col=mincol,
                                   max_col=maxcol, max_row=maxrow):
        results.append([el.value for el in row])
    return results


def getlabeledvalues(worksheet, cellrange, labelcolor=0,
                     valuecolor=8, variablelength=False, hasunits=False):
    if isinstance(cellrange, (str, unicode)):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    # print(rangetuple)
    labelcol = min(rangetuple[0], rangetuple[2])
    if rangetuple[1] < rangetuple[3]:
        minrow = rangetuple[1]
        maxrow = rangetuple[3]
    else:
        maxrow = rangetuple[1]
        minrow = rangetuple[3]
    result = {}
    if not hasunits:
        for row in worksheet.iter_rows(min_row=minrow, min_col=labelcol,
                                       max_col=labelcol + 1, max_row=maxrow):
            if row[0].value != None and row[1].value != None:
                if variablelength:
                    if (row[0].fill.start_color.index != labelcolor or
                                row[1].fill.start_color.index != valuecolor):
                        break
                result[row[0].value] = row[1].value
    else:
        for row in worksheet.iter_rows(min_row=minrow, min_col=labelcol,
                                       max_col=labelcol + 2, max_row=maxrow):
            if row[0].value != None and row[1].value != None:
                if variablelength:
                    if (row[0].fill.start_color.index != labelcolor or
                                row[1].fill.start_color.index != valuecolor):
                        break
                if row[2].value != None:
                    key = row[0].value + (' (%s)' % row[2].value)
                    result[key] = row[1].value
                else:
                    result[row[0].value] = row[1].value
    return result


def getlist(worksheet, cellrange, variablelength=False, fillcolor=8):
    if isinstance(cellrange, (str, unicode)):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    diff = (rangetuple[2] - rangetuple[0],
            rangetuple[3] - rangetuple[1])
    result = []
    if diff[0] == 0:
        listcol = rangetuple[0]
        for row in worksheet.iter_rows(min_row=rangetuple[1], min_col=listcol,
                                       max_col=listcol, max_row=rangetuple[3]):
            if variablelength:
                if not row[0].value or row[0].fill.start_color.index != fillcolor:
                    break
            result.append(row[0].value)
    elif diff[1] == 0:
        listrow = rangetuple[1]
        for col in worksheet.iter_cols(min_col=rangetuple[0], min_row=listrow,
                                       max_row=listrow, max_col=rangetuple[2]):
            if variablelength:
                if not col[0].value or col[0].fill.start_color.index != fillcolor:
                    break
            result.append(col[0])
    return result


def getinfo(worksheet, cellrange, variablelength=False, fillcolor=8,
            labels=None, inrows=True, keepempty=False):
    if isinstance(cellrange, (str, unicode)):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    result = {}
    if inrows:
        listcol = rangetuple[0]
        for row in worksheet.iter_rows(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_col=rangetuple[2], max_row=rangetuple[3]):
            if variablelength:
                if (not row[0].value or row[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not row[0].value:
                    continue
            data = [el.value for el in row[1:]]
            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                data = dict(zip(labels[1:], data))
            result[row[0].value] = data
    else:
        listrow = rangetuple[1]
        for col in worksheet.iter_cols(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_row=rangetuple[3], max_col=rangetuple[2]):
            if variablelength:
                if (not col[0].value or col[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not col[0].value:
                    continue
            data = [el.value for el in col[1:]]
            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                data = dict(zip(labels[1:], data))
            result[col[0].value] = data
    return result


def getlistinfo(worksheet, cellrange, variablelength=False, fillcolor=8,
                labels=None, inrows=True, keepempty=False):
    if isinstance(cellrange, (str, unicode)):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    # print(rangetuple)
    result = []
    if inrows:
        listcol = rangetuple[0]
        for row in worksheet.iter_rows(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_col=rangetuple[2], max_row=rangetuple[3]):
            # print(row)
            if variablelength:
                if (not row[0].value or row[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not row[0].value:
                    continue
            data = [el.value for el in row]
            # print(data)
            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                # Have to handle None in the labels
                out = {}
                for i in range(len(labels)):
                    if labels[i] == None or data[i] == None:
                        continue
                    out[labels[i]] = data[i]
                data = out
                # data = dict(zip(labels,data))
            result.append(data)
    else:
        listrow = rangetuple[1]
        for col in worksheet.iter_cols(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_row=rangetuple[3], max_col=rangetuple[2]):
            if variablelength:
                if (not col[0].value or col[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not col[0].value:
                    continue
            data = [el.value for el in col[1:]]
            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                # Have to handle None in the labels
                out = {}
                for i in range(len(labels)):
                    if labels[i] == None or data[i] == None:
                        continue
                    out[labels[i]] = data[i]
                data = out
                # data = dict(zip(labels,data))
            result.append(data)
    return result


def gettabular(worksheet, mincol, minrow, maxcol, maxrow):
    results = []
    for row in worksheet.iter_rows(min_row=minrow, min_col=mincol,
                                   max_col=maxcol, max_row=maxrow):
        results.append([el.value for el in row])
    return results


def getcellrange(worksheet, cellrange):
    try:
        mincol, minrow, maxcol, maxrow = openpyxl.utils.range_boundaries(cellrange)
    except TypeError:
        return None
    return gettabular(worksheet, mincol, minrow, maxcol, maxrow)


def scanRowForEmpty(worksheet, col, row, maxcol=256):
    count = 0
    for col in worksheet.iter_cols(min_col=col, min_row=row,
                                   max_row=row):
        if not col[0].value:
            return count
        count += 1


def scanForExpandableColumnTable(worksheet, mincol=1, minrow=1, maxrow=1,
                                 minentries=1):
    result = []
    for col in worksheet.iter_cols(min_col=mincol, min_row=minrow,
                                   max_row=maxrow):
        count = 0
        data = []
        for el in col:
            data.append(el.value)
            if el.value:
                count += 1
        if count < minentries:
            return result
        result.append(data)
    return result


def scanForHeaderRow(worksheet, mincol, minrow, header):
    maxcol = mincol + len(header) - 1
    count = 0
    for row in worksheet.iter_rows(min_col=mincol, min_row=minrow,
                                   max_col=maxcol):
        data = [el.value for el in row]
        if data == header:
            return minrow + count
        count += 1
    raise ScanFailure('Failed to find header')


def scan_for_cell_value(worksheet, mincol=None, minrow=None, maxcol=None,
                        maxrow=None, value=None):
    for row in worksheet.iter_rows(min_col=mincol, min_row=minrow,
                                   max_col=maxcol, max_row=maxrow):
        for el in row:
            if el.value == value:
                return tuple_from_coordinate(el.coordinate)
    raise ScanFailure('Failed to find cell value')


def process_zip(pc):
    separators = ['-', ' ']
    for sep in separators:
        five, guido, four = pc.partition('-')
        five = five.strip()
        four = four.strip()
        if len(five) != 5 or len(four) != 4:
            continue
        # Check for numbers?
        return five, four
    return pc, None


def prettystring(element):
    return minidom.parseString(et.tostring(element, encoding='utf-8')).toprettyxml(indent='\t',
                                                                                   encoding='utf-8')


def easymap(dictionary, inkey, outkey, parent, f=lambda x: x):
    if inkey in dictionary:
        el = et.SubElement(parent, outkey)
        el.text = f(dictionary[inkey])


def addel(outkey, parent, value):
    el = et.SubElement(parent, outkey)
    el.text = value


def addudf(parent, key, value, create=True):
    udfs = parent.find('UserDefinedFields')
    if not udfs:
        if not create:
            return
        udfs = et.SubElement(parent, 'UserDefinedFields')
    udf = et.SubElement(udfs, 'UserDefinedField')
    el = et.SubElement(udf, 'FieldName')
    el.text = key
    el = et.SubElement(udf, 'FieldValue')
    el.text = value


def easymapudf(dictionary, inkey, outkey, parent, f=lambda x: x):
    if inkey in dictionary:
        addudf(parent, outkey, f(dictionary[inkey]))


# else:
#        print(inkey)
#        raise 'WhatWhat?'

def yn2tf(s):
    return {'Y': 'true', 'N': 'false'}[s]


def repercentage(s):
    return str(s * 100) + '%'


def map_to_buildingsync(obj):
    #
    # All - Building
    #
    allbuilding = obj['All - Building']
    # Give the address
    address = et.Element('Address')
    if 'Street*' in allbuilding:
        el = et.SubElement(address, 'StreetAddressDetail')
        el = et.SubElement(el, 'Simplified')
        el = et.SubElement(el, 'StreetAddress')
        el.text = allbuilding['Street*']
    easymap(allbuilding, 'City*', 'City', address)
    easymap(allbuilding, 'State*', 'State', address)
    if 'Postal Code*' in allbuilding:
        postalcode = allbuilding['Postal Code*']
        postalcode, plus4 = process_zip(postalcode)
        postalcodeplus4 = postalcode
        if plus4:
            postalcodeplus4 += '-' + plus4
        el = et.SubElement(address, 'PostalCode')
        el.text = postalcode
        el = et.SubElement(address, 'PostalCodePlus4')
        el.text = postalcodeplus4
    # street address, city, state, zip5, zip5-4
    if len(address) == 0:
        address = None
    # Create contacts if they are present
    contacts = et.Element('Contacts')
    auditor = None
    if 'Energy Auditor' in allbuilding:
        auditor = et.SubElement(contacts, 'Contact')
        auditor.attrib['ID'] = 'EnergyAuditor'
        addel('ContactRole', auditor, 'Energy Auditor')
        addel('ContactName', auditor, allbuilding['Energy Auditor'])
        # Id, role, name
        # contacts.append(contactidXml % ('EnergyAuditor',
        #                                'Energy Auditor',
        #                                allbuilding['Energy Auditor']))
    keycontact = None
    if 'Key Contact' in allbuilding:
        keycontact = et.SubElement(contacts, 'Contact')
        keycontact.attrib['ID'] = 'KeyContact'
        addel('ContactRole', keycontact, 'Other')
        addel('ContactName', keycontact, allbuilding['Key Contact'])
        addudf(keycontact, 'ASHRAE Standard 211 Role', 'Key Contact')
        # primarycontact = 'KeyContact'
        # Id, name, role
        # contacts.append(contactidXmlCustom % ('KeyContact',
        #                                      allbuilding['Key Contact'],
        #                                      'Key Contact'))
    if 'Client Name' in allbuilding:
        client = et.SubElement(contacts, 'Contact')
        client.attrib['ID'] = 'Client'
        addel('ContactRole', client, 'Other')
        addel('ContactName', client, allbuilding['Client Name'])
        addudf(client, 'ASHRAE Standard 211 Role', 'Client')
        # Id, name, role
        # contacts.append(contactidXmlCustom % ('ClientName',
        #                                      allbuilding['Client Name'],
        #                                      'Client'))
    if 'Building Owner' in allbuilding:
        owner = et.SubElement(contacts, 'Contact')
        owner.attrib['ID'] = 'BuildingOwner'
        addel('ContactRole', owner, 'Other')
        addel('ContactName', owner, allbuilding['Building Owner'])
        addudf(owner, 'ASHRAE Standard 211 Role', 'Owner')
        # Id, name, role
        # contacts.append(contactidXmlCustom % ('BuildingOwner',
        #                                      allbuilding['Building Owner'],
        #                                      'Building Owner'))

    facilities = et.Element('Facilities')
    facility = et.SubElement(facilities, 'Facility')
    facility.attrib['ID'] = 'Building'

    easymap(allbuilding, 'Building Name*', 'PremisesName', facility)
    easymap(allbuilding, 'Facility Description - Notable Conditions',
            'PremisesNotes', facility)
    # OccupancyClassification should go here, but it can't: the enums don't match
    if 'Occupancy' in allbuilding:
        occupancy = allbuilding['Occupancy']
        if 'Typical number of occupants (during occ hours)' in occupancy:
            levels = et.SubElement(facility, 'OccupancyLevels')
            level = et.SubElement(levels, 'OccupancyLevel')
            addel('OccupantQuantity', level,
                  str(occupancy['Typical number of occupants (during occ hours)']))
        typicalocc = et.Element('TypicalOccupantUsages')
        if 'Typical occupancy (hours/week)' in occupancy:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(occupancy['Typical occupancy (hours/week)']))
            addel('TypicalOccupantUsageUnits', occ, 'Hours per week')
        if 'Typical occupancy (weeks/year)' in occupancy:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(occupancy['Typical occupancy (weeks/year)']))
            addel('TypicalOccupantUsageUnits', occ, 'Weeks per year')
        if len(typicalocc) > 0:
            facility.append(typicalocc)
        if 'Number of Dwelling Units in Building (Multifamily Only)' in occupancy:
            units = et.SubElement(facility, 'SpatialUnits')
            addel('SpatialUnitType', units, 'Apartment units')
            addel('NumberOfUnits', units,
                  str(occupancy['Number of Dwelling Units in Building (Multifamily Only)']))

    easymap(allbuilding, 'Conditioned Floors Above grade',
            'ConditionedFloorsAboveGrade', facility, f=str)
    easymap(allbuilding, 'Conditioned Floors Below grade',
            'ConditionedFloorsBelowGrade', facility, f=str)
    easymap(allbuilding, 'Building automation system? (Y/N)',
            'BuildingAutomationSystem', facility, yn2tf)
    easymap(allbuilding, 'Historical landmark status? (Y/N)',
            'HistoricalLandmark', facility, yn2tf)
    # Map to FloorAreas
    floorareas = et.Element('FloorAreas')
    if 'Total conditioned area' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Conditioned')
        addel('FloorAreaValue', floorarea, allbuilding['Total conditioned area'])
    if 'Gross floor area' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Gross')
        addel('FloorAreaValue', floorarea, allbuilding['Gross floor area'])
    if 'Conditioned area (heated only)' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Cooled only')
        addel('FloorAreaValue', floorarea, allbuilding['Conditioned area (heated only)'])
    if 'Conditioned area (cooled only)' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Heated only')
        addel('FloorAreaValue', floorarea, allbuilding['Conditioned area (cooled only)'])
    # Map Space Function table to FloorAreas
    if 'Space Function' in allbuilding:
        for key, value in allbuilding['Space Function'].items():
            floorarea = et.SubElement(floorareas, 'FloorArea')
            addel('FloorAreaType', floorarea, 'Custom')
            addel('FloorAreaCustomName', floorarea, key)
            addel('FloorAreaValue', floorarea, value)

    easymap(allbuilding, 'Year of construction*',
            'YearOfConstruction', facility, f=str)

    easymap(allbuilding, 'Year of Prior Energy Audit',
            'YearOfLastEnergyAudit', facility, f=str)

    easymap(allbuilding, 'Last Renovation*',
            'YearOfLastMajorRemodel', facility, f=str)
    #
    # All - Space Functions
    #
    subsections = et.Element('Subsections')
    spacefunctions = obj['All - Space Functions']
    for key, value in spacefunctions.items():
        subsection = et.SubElement(subsections, 'Subsection')
        # First the stuff that has a slot to go into
        addel('PremisesName', subsection, key)
        if 'Number of Occupants' in value:
            levels = et.SubElement(subsection, 'OccupancyLevels')
            level = et.SubElement(levels, 'OccupancyLevel')
            addel('OccupantQuantity', level,
                  str(value['Number of Occupants']))
        typicalocc = et.Element('TypicalOccupantUsages')
        if 'Use (hours/week)' in value:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(value['Use (hours/week)']))
            addel('TypicalOccupantUsageUnits', occ, 'Hours per week')
        if 'Use (weeks/year)' in value:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(value['Use (weeks/year)']))
            addel('TypicalOccupantUsageUnits', occ, 'Weeks per year')
        if len(typicalocc) > 0:
            subsection.append(typicalocc)
        if 'Gross Floor Area' in value:
            floorareas = et.SubElement(subsection, 'FloorAreas')
            floorarea = et.SubElement(floorareas, 'FloorArea')
            addel('FloorAreaType', floorarea, 'Gross')
            addel('FloorAreaValue', floorarea, str(value['Gross Floor Area']))
        # Now for the UDFs
        easymapudf(value, 'Function type',
                   'ASHRAE Standard 211 Function Type', subsection)
        easymapudf(value, 'Original intended use',
                   'ASHRAE Standard 211 Original Intended Use', subsection)
        easymapudf(value, 'Percent Conditioned Area',
                   'ASHRAE Standard 211 Percent Conditioned Area', subsection,
                   f=repercentage)
        easymapudf(value, 'Approximate Plug Loads (W/sf)',
                   'ASHRAE Standard 211 Approximate Plug Loads', subsection, f=str)
        easymapudf(value, 'Principal HVAC Type',
                   'ASHRAE Standard 211 Principal HVAC Type', subsection, f=str)
        easymapudf(value, 'Principal Lighting Type',
                   'ASHRAE Standard 211 Principal Lighting Type', subsection, f=str)

    if len(subsections) > 0:
        facility.append(subsections)

    # Now for the UDFs from All - Building
    easymapudf(allbuilding, 'Primary Building use type*',
               'ASHRAE Standard 211 Primary Building Use Type', facility)
    easymapudf(allbuilding, 'Year Last Commissioned',
               'ASHRAE Standard 211 Year Last Commissioned', facility, f=str)
    easymapudf(allbuilding, 'Percent owned (%)',
               'ASHRAE Standard 211 Percent Owned', facility, f=repercentage)
    easymapudf(allbuilding, 'Percent leased (%)',
               'ASHRAE Standard 211 Percent Leased', facility, f=repercentage)
    easymapudf(allbuilding, 'Total Number of Floors',
               'ASHRAE Standard 211 Total Number of Floors', facility, f=str)
    if 'Excluded Spaces' in allbuilding:
        allbuilding['Excluded Spaces'] = ', '.join(allbuilding['Excluded Spaces'])
    easymapudf(allbuilding, 'Excluded Spaces',
               'ASHRAE Standard 211 Excluded Spaces', facility)

    if 'Occupancy' in allbuilding:
        easymapudf(allbuilding['Occupancy'],
                   '% of Dwelling Units currently Occupied (Multifamily Only)',
                   'ASHRAE Standard 211 Percent Dwelling Units Currently Occupied',
                   facility, f=repercentage)

    # Wrap up for facility
    if len(facility) == 0:
        facility = None
        facilities = None

    # Map energy sources to a report
    report = et.Element('Report')

    if 'Energy Sources' in allbuilding:
        scenarios = et.SubElement(report, 'Scenarios')
        scenario = et.SubElement(scenarios, 'Scenario')
        scenario.attrib['ID'] = 'ASHRAEStandard211EnergySources'
        addel('ScenarioName', scenario, 'ASHRAE Standard 211 Energy Sources')
        resources = et.SubElement(scenario, 'ResourceUses')
        for el in allbuilding['Energy Sources']:
            resource = et.Element('ResourceUse')
            # Nope, enum fail on both
            # easymap(el, 'Energy Source', 'EnergyResource', resource)
            # if 'Type' in el:
            #    sub = et.SubElement(resource, 'Utility')
            #    sub = et.SubElement(sub, 'MeteringConfiguration')
            #    sub.text = el['Type']
            easymapudf(el, 'Energy Source', 'ASHRAE Standard 211 Energy Source',
                       resource)
            easymapudf(el, 'Type', 'ASHRAE Standard 211 Type', resource)
            easymapudf(el, 'ID', 'ASHRAE Standard 211 ID', resource, f=str)
            easymapudf(el, 'Rate schedule', 'ASHRAE Standard 211 Rate Schedule',
                       resource, f=str)
            if len(resource) > 0:
                resources.append(resource)
        if len(scenario) > 0 and facility:
            link = et.SubElement(scenario, 'LinkedPremises')
            el = et.SubElement(link, 'Facility')
            el = et.SubElement(el, 'LinkedFacilityID')
            el.attrib['IDref'] = facility.attrib['ID']
    if auditor:
        el = et.SubElement(report, 'AuditorContactID')
        el.attrib['IDref'] = auditor.attrib['ID']

    easymapudf(allbuilding, 'Date of site visit(s)',
               'ASHRAE Standard 211 Date of site visit(s)', report)

    # Wrap up for report
    if len(report) == 0:
        report = None
    #
    # L1 - EEM Summary
    #
    fields = ['Modified System',
              'Impact on Occupant Comfort or IEQ',
              'Other Non-Energy Impacts',
              'Cost',
              'Savings Impact',
              'Typical ROI',
              'Priority']
    # First the low cost items
    summary = obj['L1 - EEM Summary']
    measures = et.Element('Measures')
    if 'Low-Cost and No-Cost Recommendations' in summary:
        for key, value in summary['Low-Cost and No-Cost Recommendations'].items():
            # print(key)
            measure = et.SubElement(measures, 'Measure')
            el = et.SubElement(measure, 'LongDescription')
            el.text = key
            udfs = et.SubElement(measure, 'UserDefinedFields')
            for field in fields:
                if field in value:
                    udf = et.SubElement(udfs, 'UserDefinedField')
                    udfname = et.SubElement(udf, 'FieldName')
                    udfname.text = field
                    udfvalue = et.SubElement(udf, 'FieldValue')
                    udfvalue.text = value[field]
            udf = et.SubElement(udfs, 'UserDefinedField')
            udfname = et.SubElement(udf, 'FieldName')
            udfname.text = 'ASHRAE Standard 211 Measure Category'
            udfvalue = et.SubElement(udf, 'FieldValue')
            udfvalue.text = 'Low-Cost and No-Cost Recommendations'
    # Change that one thing...
    fields[1] = 'Impact on Occupant Comfort'
    if 'Potential Capital Recommendations' in summary:
        for key, value in summary['Potential Capital Recommendations'].items():
            # print(key)
            measure = et.SubElement(measures, 'Measure')
            el = et.SubElement(measure, 'LongDescription')
            el.text = key
            udfs = et.SubElement(measure, 'UserDefinedFields')
            for field in fields:
                if field in value:
                    udf = et.SubElement(udfs, 'UserDefinedField')
                    udfname = et.SubElement(udf, 'FieldName')
                    udfname.text = field
                    udfvalue = et.SubElement(udf, 'FieldValue')
                    udfvalue.text = value[field]
            udf = et.SubElement(udfs, 'UserDefinedField')
            udfname = et.SubElement(udf, 'FieldName')
            udfname.text = 'ASHRAE Standard 211 Measure Category'
            udfvalue = et.SubElement(udf, 'FieldValue')
            udfvalue.text = 'Potential Capital Recommendations'

    #
    # Assemble the final result
    #
    bsxml = et.Element('Audits')
    bsxml.attrib['xmlns'] = "http://nrel.gov/schemas/bedes-auc/2014"
    bsxml.attrib['xmlns:xsi'] = "http://www.w3.org/2001/XMLSchema-instance"
    bsxml.attrib[
        'xsi:schemaLocation'] = "http://nrel.gov/schemas/bedes-auc/2014 ../BuildingSync.xsd"
    audit = et.SubElement(bsxml, 'Audit')
    # First is Sites
    if address or keycontact or facilities:
        sites = et.SubElement(audit, 'Sites')
        site = et.SubElement(sites, 'Site')
        if address:
            site.append(address)
        if keycontact:
            pcid = et.SubElement(site, 'PrimaryContactID')
            pcid.text = keycontact.attrib['ID']
        if facilities:
            site.append(facilities)
    # Next is Measures
    if measures:
        audit.append(measures)
    # Now Reports
    if report:
        audit.append(report)
    # Last is Contacts
    if contacts:
        audit.append(contacts)
    # Done!
    return bsxml


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Convert ASHRAE Std. 211 Workbook into BSXML.')
    parser.add_argument('infile', metavar='INFILE')
    parser.add_argument('-p, --pretty', dest='pretty', action='store_true',
                        help='output pretty xml')
    parser.add_argument('-o, --output', dest='outfile', action='store',
                        default='std211.xml',
                        help='file to save BSXML output in')

    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.infile)
    std211 = read_std211_xls(wb)
    bsxml = map_to_buildingsync(std211)
    print(prettystring(bsxml).decode('utf-8'))
    fp = open(args.outfile, 'w')
    if args.pretty:
        fp.write(prettystring(bsxml).decode('utf-8'))
    else:
        fp.write('<?xml version="1.0" encoding="UTF-8"?>')
        fp.write(et.tostring(bsxml, encoding='utf-8').decode('utf-8'))
    fp.close()
